import openpyxl as xl
import datetime as dt
import names
from random import choice, randint

""" 
Populate the sample spreadsheet with mock data.
"""


def populateSheet(rows):
    # Load the empty sample spreadsheet

    try:
        wb = xl.load_workbook(filename='Workbook 1.xlsx')
        ws = wb['Spreadsheet 1']
    except:
        print('File name or spreadsheet name error.')
        quit()
    
    course = {'APDE 1000': 'Ideas and Imagery', 'APDE 1001': 'Design Fundamentals', 'APDE 1002': 'Colour and Design', 'DRWG 1004': 'Fundamentals of Drawing', 'APDE 1006': 'Understanding Art and Design', 'FINA 1002': 'Fine Art Studio',
              'APDE 1005': 'Visual Art Studio: 3D', 'GRPH 1003': 'Graphic Design Studio', 'PHOT 1006': 'Introduction to Digital Photography', 'DRWG 1007': 'Drawing From Your Imagination', 'DRWG 1005': 'Analytical Drawing', 'DRWG 1006': 'Life Drawing'}

    # Erase the existing data
    while ws.max_row > 1:
        for i in range(2, ws.max_row + 1):
            ws.delete_rows(i)

    # Populate the spreadsheet

    ids = []

    while ws.max_row != rows + 1:
        for i in range(rows): # Loop for number of rows of times
            # Academic Period
            ws['A' + str(i + 2)] = choice(('2023', '2024')) + choice((' Winter', ' Summer', ' Fall'))
            # Campus
            ws['B' + str(i + 2)] = choice(('Barrie', 'Midland', 'Orangeville', 'Orillia'))
            # Delivery
            ws['C' + str(i + 2)] = 'GC Flex'        
            # CRN
            ws['G' + str(i + 2)] = randint(10000, 99999)
            # Course & Course Title
            ws['H' + str(i + 2)] = choice(tuple(course.keys()))
            ws['i' + str(i + 2)] = course[ws['H' + str(i + 2)].value]
            # Building
            ws['K' + str(i + 2)] = choice(('A Building', 'B Building', 'C Building', 'D Building', 'E Building'))
            # Room
            ws['L' + str(i + 2)] = str(randint(101, 335))
            # Date
            ws['N' + str(i + 2)] = choice(('2023-05-08', '2023-05-09', '2023-05-10', '2023-05-11', '2023-05-12'))
            ws['O' + str(i + 2)] = (dt.date.fromisoformat(ws['N' + str(i + 2)].value) + dt.timedelta(days=98)).isoformat()
            ws['P' + str(i + 2)] = 98
            # Time
            ws['Q' + str(i + 2)] = (dt.datetime(int(ws['A' + str(i + 2)].value[:4]), int(ws['N' + str(i + 2)].value[5:7]), int(ws['N' + str(i + 2)].value[8:10]), hour=(choice((7, 8, 9, 10, 14, 15, 16, 18, 19))))).strftime("%H:%M")
            ws['R' + str(i + 2)] = (dt.datetime.strptime(ws['Q' + str(i + 2)].value, "%H:%M") + dt.timedelta(hours=2, minutes=50)).strftime("%H:%M")
            # Actual Enrolment
            ws['S' + str(i + 2)] = randint(20, 40)
            # Room Capacity
            ws['T' + str(i + 2)] = 50
            # Instructor ID
            while True:
                id = randint(100000, 999999)
                if id not in ids:
                    ids.append(id)
                    break
            ws['U' + str(i + 2)] = id
            # Instructor First Name
            ws['V' + str(i + 2)] = names.get_first_name()
            # Instructor Last Name
            ws['W' + str(i + 2)] = names.get_last_name()
            # Instructor Email
            ws['X' + str(i + 2)] = ws['V' + str(i + 2)].value + '.' + ws['W' + str(i + 2)].value + '@georgiancollege.com'

    # Style the text in the cells
    for i in range(2, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).alignment = xl.styles.Alignment(horizontal='left', vertical='center')
            # Set the font famaily and size
            ws.cell(row=i, column=j).font = xl.styles.Font(name='IBM Plex Sans', size=10.5)
            # Set the border
            ws.cell(row=i, column=j).border = xl.styles.Border(left=xl.styles.Side(border_style='thin', color='FF000000'), right=xl.styles.Side(border_style='thin', color='FF000000'), top=xl.styles.Side(border_style='thin', color='FF000000'), bottom=xl.styles.Side(border_style='thin', color='FF000000'))
    
    # Save the populated spreadsheet
    print(str(ws.max_row - 1) + ' rows have been populated.')
    wb.save('Workbook 1.xlsx')

    # Copy the spreadsheet to a new workbook
    wb2 = xl.Workbook()
    ws2 = wb2.active
    for row in ws.rows:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    # Delete random rows from Spreadsheet 2
    for i in range(randint(1, rows)):
        ws2.delete_rows(randint(2, ws2.max_row))

    # Style the text in the cells
    for i in range(2, ws2.max_row + 1):
        for j in range(1, ws2.max_column + 1):
            ws2.cell(row=i, column=j).alignment = xl.styles.Alignment(horizontal='left', vertical='center')
            # Set the font famaily and size
            ws2.cell(row=i, column=j).font = xl.styles.Font(name='IBM Plex Sans', size=10.5)
            # Set the border
            ws2.cell(row=i, column=j).border = xl.styles.Border(left=xl.styles.Side(border_style='thin', color='FF000000'), right=xl.styles.Side(border_style='thin', color='FF000000'), top=xl.styles.Side(border_style='thin', color='FF000000'), bottom=xl.styles.Side(border_style='thin', color='FF000000'))
    
    # Save Workbook 2
    print(str(ws.max_row - ws2.max_row) + ' rows deleted in Workbook 2.')
    wb2.save('Workbook 2.xlsx')

populateSheet(randint(300, 500))