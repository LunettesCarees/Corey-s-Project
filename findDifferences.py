import openpyxl as xl

"""
Compare two Excel files and return the differences in the output file.
"""

def findDifferences(file1, file2):
    # Load the spreadsheets
    wb1 = xl.load_workbook(file1)
    ws1 = wb1.active
    wb2 = xl.load_workbook(file2)
    ws2 = wb2.active

    # Get all the instructor IDs from Spreadsheet 1
    ids1 = []
    for i in range(2, ws1.max_row + 1):
        ids1.append(ws1['U' + str(i)].value)

    # Get all the instructor IDs from Spreadsheet 2
    ids2 = []
    for i in range(2, ws2.max_row + 1):
        ids2.append(ws2['U' + str(i)].value)

    if len(ids1) == len(ids2):
        print('There are no differences between the two spreadsheets.')
        return
    
    if len(ids1) < len(ids2):
        print('The spreadsheet with all the instructors has less rows than the spreadsheet with the experienced instructors.\nPlease check the spreadsheets and try again.')
        return

    # Get the differences between the two spreadsheets
    differences = []
    for i in range(len(ids1)):
        if ids1[i] not in ids2:
            differences.append(ids1[i])

    # Load the output spreadsheet
    wb3 = xl.Workbook()
    ws3 = wb3.active
    ws3['A1'] = 'Instructor ID'
    ws3['B1'] = 'Instructor First Name'
    ws3['C1'] = 'Instructor Last Name'
    ws3['D1'] = 'Instructor Email'

    # Write the differences to the output spreadsheet
    for i in range(len(differences)):
        for j in range(2, ws1.max_row + 1):
            if ws1['U' + str(j)].value == differences[i]:
                ws3['A' + str(i + 2)] = ws1['U' + str(j)].value
                ws3['B' + str(i + 2)] = ws1['V' + str(j)].value
                ws3['C' + str(i + 2)] = ws1['W' + str(j)].value
                ws3['D' + str(i + 2)] = ws1['X' + str(j)].value

    # Save the output spreadsheet
    wb3.save('output.xlsx')
    print('The output spreadsheet has been saved as output.xlsx.')